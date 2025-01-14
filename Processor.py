
#### Process ####
# Pull information from an excel sheet
## Get correct cell collumn x

test_workbook = 'Demo/Demo Sheet.xlsx'
config_workbook = 'Config/Config.xlsx'
print_preview = 'view.html'

import openpyxl

# Pull information from Demo Resource Sheet, This has the device names
source_book = openpyxl.load_workbook(test_workbook, data_only=True, read_only=True)
#print(workbook.sheetnames)
sheet = source_book['MASTER WORKSHEET']
devices = [] # Create list "devices"
for i, column in enumerate(sheet):
        if i <= 1: # Skip number of rows
                continue
        device = column[0].value # get value from column number, starting at 0
        devices.append(device) # add to list "devices"
print (devices)

# Pull information from Config Workbook, this has the device types and device tests
config_book = openpyxl.load_workbook(config_workbook, data_only=True, read_only=True)
#print(config_book.sheetnames)
device_sheet = config_book['Device Types']
device_types = [] 
for i, column in enumerate(device_sheet):
        if i <= 0: 
                continue
        device_type = column[0].value 
        device_types.append(device_type) 
print (device_types)

test_sheet = config_book['Test Types']
test_types = [] 
for i, column in enumerate(test_sheet):
        if i <= 0:
                continue
        test_type = column[0].value 
        test_types.append(test_type) 
print (test_types)


# SAMPLE #
# Define the list of tests with alt text
tests_with_alt_text = {
    "test1": "on/off",
    "test2": "short",
    "test3": "open",
    "test4": "close"
}

# Define the table of device types and their corresponding tests
device_tests = {
    "cam": ["test1", "test2"],
    "kpd": ["test1", "test2", "test3", "test4"]
}

# Function to get the tests for each device with alt text
def get_device_tests_with_alt_text(device_names):
    result = {}
    for device in device_names:
        # Extract the device type (e.g., 'cam' from 'cam1')
        device_type = ''.join([i for i in device if not i.isdigit()])
        # Get the corresponding tests for the device type
        tests = device_tests.get(device_type, [])
        # Add alt text to each test
        result[device] = [(test, tests_with_alt_text[test]) for test in tests]
    return result

# List of device names
device_names = ["cam1", "kpd2", "kpd5", "cam4"]

# Get the tests for each device with alt text and print the results
device_tests_result = get_device_tests_with_alt_text(device_names)
print(device_tests_result)



####

# Label information by filter
## Create a list of POSSIBLE_VALUES that SAMPLE_VALUE can have
## Create a list of OUTPUT_VALUES
## Create an association of POSSIBLE_VALUES with OUTPUT VALUES
## Relate OUTPUT_VALUE to SAMPLE_VALUE

## IF SAMPLE_VALUE = A; THEN OUTPUT_VALUE = 1,2,3
## IF SAMPLE_VALUE = B; THEN OUTPUT_VALUE = 2,4,6


# Create a template for each label
## Define template

# Add values to a list



# Set up HTML page for print
## these are items that create the print preview
## Add associated listed items into print preview

### Test Template Parts
### This is all HTML garbage that will eventually go into the format of the target page.
### This really should have been done in JQuery, but that would require pulling an external library
Device_Header = '<h3>DEVICE HEADER</h3>'
Test_Header = '<center><table><tr><th>TESTNAME</th> <th>Yes</th> <th>No</th> <th>N/A</th> <th>Comment</th></tr>'
Test_Line = '<td>Test Name Variable Goes Here</td><td><center><input type="checkbox"></center></td> <td><center><input type="checkbox"></center></td> <td><center><input type="checkbox"></center></td> <td id=testCommentCell></td></tr>'
Test_Close = '</table></center>'


html_open = '<!DOCTYPE html><html><head><link rel="stylesheet" href="style.css"></head><body>'
header_block = '<head><title>Page Title</title></head><h1>Header</h1><br><left><table><tr><td><h2>Name:</h2><h2>Location:</h2></td><td></td><td><h2>Authorization:</h2></td><td></td></table><br>'
html_close = '</body></html>'

signature_block = '<center><table><tr><td><h2>Approver</h2></td><td id=testCommentCell;></td></tr><tr><td><h2>Signature</h2></td><td id=testCommentCell;></td></tr></table></center>'
comment_block = '<br><br> Comments: <center><table><tr><td id=testCommentCell></td></tr><tr><td id=testCommentCell></td></tr><tr><td id=testCommentCell></td></tr><tr><td id=testCommentCell></td></tr><tr><td>  </td></tr></table></center>'

# Define Print Preview File and Clear Print Preview

file = open(print_preview, 'w')
file.write("")
file.close()
# Open Print Preview to be appended
file = open(print_preview, 'a')
file.write(html_open)

### Top of page junk
file.write(header_block)


### Associate the variables from the demo data with the tests from the config


### Dummy output for a device
file.write(Device_Header)
file.write(Test_Header)
file.write(Test_Line)
file.write(Test_Line)
file.write(Test_Line)
file.write(Test_Line)
file.write(Test_Line)
file.write(Test_Close)

#for each TEMPLATE_ITEM in QUEUE
#    echo TEMPLATE_ITEM >> view.html


file.write(comment_block)
file.write(signature_block)
file.write(html_close)
file.close()


# Create a pdf from that information
## send to pdf printer

#### Interface ####
# Create user interface
## Create interface window
## Create filepath selection field
## Create browse button
## Create sort dropdown
## Creat "Run" button