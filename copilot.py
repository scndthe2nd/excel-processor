import openpyxl

# Define the file paths
test_workbook = 'Demo/Demo Sheet.xlsx'
config_workbook = 'Config/Config.xlsx'
print_preview = 'view.html'

# Load the config_workbook
config_wb = openpyxl.load_workbook(config_workbook)

# Load the Device Types worksheet
device_types_ws = config_wb['Device Types']

# Load the Test Types worksheet
test_types_ws = config_wb['Test Types']

# Create a dictionary for device types and their corresponding tests
device_tests = {}
for row in device_types_ws.iter_rows(min_row=2, values_only=True):
    device_type, tests = row
    device_tests[device_type] = tests.split(',')

# Create a dictionary for tests and their alt text
tests_with_alt_text = {}
for row in test_types_ws.iter_rows(min_row=2, values_only=True):
    test, alt_text = row
    tests_with_alt_text[test] = alt_text

# Load the source_workbook
source_wb = openpyxl.load_workbook(test_workbook)

# Load the Master Worksheet
master_ws = source_wb['MASTER WORKSHEET']

# Get the list of device names from the Master Worksheet
device_names = [row[0] for row in master_ws.iter_rows(min_row=2, values_only=True)]

# Function to get the tests for each device with alt text
def get_device_tests_with_alt_text(device_names):
    result = {}
    for device in device_names:
        # Extract the device type (e.g., 'cam' from 'cam1')
        device_type = ''.join([i for i in device if not i.isdigit()])
        # Get the corresponding tests for the device type
        tests = device_tests.get(device_type, [])
        # Add alt text to each test
        result[device] = [(test, tests_with_alt_text[test]) for test in tests]
    return result

# Get the tests for each device with alt text
device_tests_result = get_device_tests_with_alt_text(device_names)

# Generate HTML content
html_content = """
<!DOCTYPE html>
<html>
<head>
    <title>Device Tests</title>
</head>
<body>
    <h1>Device Tests</h1>
    <table border="1">
        <tr>
            <th>Device Name</th>
            <th>Tests</th>
        </tr>
"""

for device, tests in device_tests_result.items():
    html_content += f"<tr><td>{device}</td><td>"
    for test, alt_text in tests:
        html_content += f"{test} ({alt_text})<br>"
    html_content += "</td></tr>"

html_content += """
    </table>
</body>
</html>
"""

# Save HTML content to a file
with open(print_preview, "w") as file:
    file.write(html_content)

print(f"HTML file '{print_preview}' has been created.")